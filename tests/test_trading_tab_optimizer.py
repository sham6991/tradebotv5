import os
import tempfile
import unittest
from unittest import mock

import openpyxl

from settings_service import DEFAULT_SETTINGS
from tests.test_backtest_export import nifty_frame, option_frame
from trade_settings_optimizer import OptimizerStopped
from trading_tab_optimizer import FIXED_TRADING_TAB_KEYS, TRADING_TAB_OPTIMIZED_KEYS, run_trading_tab_optimizer
from web_app import settings_from_values


class TradingTabOptimizerTests(unittest.TestCase):
    def test_optimizer_exports_trading_tab_settings_without_changing_fixed_live_keys(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp_dir:
            settings = settings_from_values({
                **DEFAULT_SETTINGS,
                "max_trades": "2",
                "buy_limit_score_low": "40",
                "market_entry_score": "60",
                "fast_ohlcv_entry_enabled": "true",
                "backtest_limit_fill_mode": "CONSERVATIVE",
                "aggressive_live_entry_enabled": "false",
                "one_entry_attempt_per_candle": "true",
                "max_spread_points": "2",
            })
            fixed_fast_entry = settings["fast_ohlcv_entry_enabled"]
            fixed_limit_fill_mode = settings["backtest_limit_fill_mode"]
            progress = []

            result = run_trading_tab_optimizer(
                nifty_frame(count=12),
                [option_frame(count=12), option_frame(count=12)],
                settings,
                temp_dir,
                source_metadata={"data_source_label": "unit test"},
                optimizer_config={
                    "candidate_ranges": {
                        "buy_limit_score_low": [35, 40],
                        "market_entry_score": [50, 60],
                        "minimum_body_percent": [20],
                        "minimum_close_position": [55],
                        "market_entry_minimum_body_percent": [25],
                        "market_entry_minimum_close_position": [60],
                        "trigger_upper_wick_max": [45],
                        "hard_rejection_upper_wick_max": [50],
                        "volume_previous_multiplier": [0.8],
                        "avg_volume_minimum_multiplier": [0.5],
                        "volume_pickup_avg_multiplier": [0.7],
                        "large_candle_multiplier": [2.2],
                        "move_from_low_max_multiplier": [1.1],
                        "gap_spike_multiplier": [1.2],
                        "buy_limit_offset_multiplier": [0.15],
                        "minimum_offset": [1],
                        "maximum_offset": [2],
                        "enable_chop_filter": ["Disabled"],
                        "chop_lookback_candles": [3],
                        "chop_overlap_count": [2],
                        "missed_limit_cooldown_candles": [0],
                    },
                    "parallel_workers": 2,
                },
                progress_callback=progress.append,
            )

            self.assertEqual(settings["fast_ohlcv_entry_enabled"], fixed_fast_entry)
            self.assertEqual(settings["backtest_limit_fill_mode"], fixed_limit_fill_mode)
            self.assertGreaterEqual(result["runs"], 4)
            self.assertEqual(result["parallel_workers"], 2)
            self.assertEqual(result["optimized_keys"], TRADING_TAB_OPTIMIZED_KEYS)
            self.assertEqual(result["fixed_trading_keys"], FIXED_TRADING_TAB_KEYS)
            self.assertTrue(result["output_path"].startswith(temp_dir))
            self.assertTrue(os.path.exists(result["output_path"]))
            self.assertIn("buy_limit_score_low", result["best_settings"])
            self.assertIn("market_entry_score", result["best_settings"])
            self.assertEqual(result["best_settings"]["fast_ohlcv_entry_enabled"], fixed_fast_entry)
            self.assertEqual(result["best_settings"]["backtest_limit_fill_mode"], fixed_limit_fill_mode)
            self.assertIn("Changed Settings", result["best_reliable_result"])
            self.assertIn("Phase Coverage %", result["best_reliable_result"])
            self.assertNotIn("Profit Points", result["best_reliable_result"])
            self.assertTrue(any(item["percent"] > 0 for item in progress))
            self.assertEqual(progress[-1]["stage"], "Finalizing")
            self.assertEqual(progress[-1]["percent"], 100)

            workbook = openpyxl.load_workbook(result["output_path"], read_only=True)
            self.assertIn("Optimizer Guide", workbook.sheetnames)
            self.assertIn("Best Trading Settings", workbook.sheetnames)
            self.assertIn("Ranked Results", workbook.sheetnames)
            self.assertIn("Phase Breakdown", workbook.sheetnames)
            self.assertIn("Candidate Ranges", workbook.sheetnames)

    def test_optimizer_skips_full_option_column_rebuild_inside_candidate_loop(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp_dir:
            settings = settings_from_values({
                **DEFAULT_SETTINGS,
                "max_trades": "2",
                "buy_limit_score_low": "40",
                "market_entry_score": "60",
            })
            with mock.patch(
                "engine.ensure_option_formula_columns",
                side_effect=AssertionError("candidate loop should not rebuild full option score columns"),
            ):
                result = run_trading_tab_optimizer(
                    nifty_frame(count=12),
                    [option_frame(count=12), option_frame(count=12)],
                    settings,
                    temp_dir,
                    optimizer_config={
                        "candidate_ranges": {
                            "buy_limit_score_low": [35, 40],
                            "market_entry_score": [50, 60],
                            "minimum_body_percent": [20],
                            "minimum_close_position": [55],
                            "market_entry_minimum_body_percent": [25],
                            "market_entry_minimum_close_position": [60],
                            "trigger_upper_wick_max": [45],
                            "hard_rejection_upper_wick_max": [50],
                            "volume_previous_multiplier": [0.8],
                            "avg_volume_minimum_multiplier": [0.5],
                            "volume_pickup_avg_multiplier": [0.7],
                            "large_candle_multiplier": [2.2],
                            "move_from_low_max_multiplier": [1.1],
                            "gap_spike_multiplier": [1.2],
                            "buy_limit_offset_multiplier": [0.15],
                            "minimum_offset": [1],
                            "maximum_offset": [2],
                            "enable_chop_filter": ["Disabled"],
                            "chop_lookback_candles": [3],
                            "chop_overlap_count": [2],
                            "missed_limit_cooldown_candles": [0],
                        },
                        "parallel_workers": 1,
                    },
                )

            self.assertGreater(result["runs"], 0)

    def test_trading_tab_optimizer_can_be_stopped_before_writing_report(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp_dir:
            settings = settings_from_values({**DEFAULT_SETTINGS, "max_trades": "2"})
            with self.assertRaises(OptimizerStopped):
                run_trading_tab_optimizer(
                    nifty_frame(count=12),
                    [option_frame(count=12), option_frame(count=12)],
                    settings,
                    temp_dir,
                    optimizer_config={
                        "candidate_ranges": {
                            "buy_limit_score_low": [35, 40],
                            "market_entry_score": [50, 60],
                            "minimum_body_percent": [20],
                            "minimum_close_position": [55],
                            "market_entry_minimum_body_percent": [25],
                            "market_entry_minimum_close_position": [60],
                            "trigger_upper_wick_max": [45],
                            "hard_rejection_upper_wick_max": [50],
                            "volume_previous_multiplier": [0.8],
                            "avg_volume_minimum_multiplier": [0.5],
                            "volume_pickup_avg_multiplier": [0.7],
                            "large_candle_multiplier": [2.2],
                            "move_from_low_max_multiplier": [1.1],
                            "gap_spike_multiplier": [1.2],
                            "buy_limit_offset_multiplier": [0.15],
                            "minimum_offset": [1],
                            "maximum_offset": [2],
                            "enable_chop_filter": ["Disabled"],
                            "chop_lookback_candles": [3],
                            "chop_overlap_count": [2],
                            "missed_limit_cooldown_candles": [0],
                        },
                        "parallel_workers": 1,
                    },
                    stop_requested=lambda: True,
                )
            self.assertEqual(os.listdir(temp_dir), [])


if __name__ == "__main__":
    unittest.main()

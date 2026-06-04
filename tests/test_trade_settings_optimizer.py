import os
import tempfile
import unittest

import openpyxl

from tests.test_backtest_export import backtest_settings, nifty_frame, option_frame
from trade_settings_optimizer import (
    OPTIMIZED_RISK_SETTING_KEYS,
    OptimizerStopped,
    optimizer_worker_count,
    run_risk_settings_optimizer,
)


class TradeSettingsOptimizerTests(unittest.TestCase):
    def test_optimizer_exports_ranked_risk_settings_without_strategy_changes(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp_dir:
            settings = backtest_settings(max_trades=2, buy_limit_score_low=40)
            original_entry_score = settings["buy_limit_score_low"]
            progress = []
            result = run_risk_settings_optimizer(
                nifty_frame(count=12),
                [option_frame(count=12), option_frame(count=12)],
                settings,
                temp_dir,
                source_metadata={"data_source_label": "unit test"},
                optimizer_config={
                    "candidate_ranges": {
                        "profit_points": [4, 8],
                        "safety_points": [4, 8],
                        "time_exit": [2],
                        "cooldown": [0, 1],
                        "max_trades": [1, 2],
                    },
                    "refine_top_n": 0,
                    "parallel_workers": 2,
                },
                progress_callback=progress.append,
            )

            self.assertEqual(settings["buy_limit_score_low"], original_entry_score)
            self.assertEqual(result["runs"], 16)
            self.assertEqual(result["parallel_workers"], 2)
            self.assertEqual(result["optimized_keys"], OPTIMIZED_RISK_SETTING_KEYS)
            self.assertTrue(result["output_path"].startswith(temp_dir))
            self.assertTrue(os.path.exists(result["output_path"]))
            self.assertIn("profit_points", result["best_settings"])
            self.assertIn("safety_points", result["best_settings"])
            self.assertEqual(result["best_settings"]["buy_limit_score_low"], original_entry_score)
            self.assertIn("Phase Coverage %", result["best_reliable_result"])
            self.assertTrue(any(item["percent"] > 0 for item in progress))
            self.assertEqual(progress[-1]["stage"], "Finalizing")
            self.assertEqual(progress[-1]["percent"], 100)

            workbook = openpyxl.load_workbook(result["output_path"], read_only=True)
            self.assertIn("Optimizer Guide", workbook.sheetnames)
            self.assertIn("Best Settings", workbook.sheetnames)
            self.assertIn("Ranked Results", workbook.sheetnames)
            self.assertIn("Phase Breakdown", workbook.sheetnames)
            self.assertIn("Candidate Ranges", workbook.sheetnames)

    def test_optimizer_worker_count_is_capped_for_small_laptops(self):
        self.assertEqual(optimizer_worker_count({"parallel_workers": 99}, 20), 2)
        self.assertEqual(optimizer_worker_count({"parallel_workers": 1}, 20), 1)
        self.assertEqual(optimizer_worker_count({"parallel_workers": 2}, 1), 1)

    def test_risk_optimizer_can_be_stopped_before_writing_report(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp_dir:
            with self.assertRaises(OptimizerStopped):
                run_risk_settings_optimizer(
                    nifty_frame(count=12),
                    [option_frame(count=12), option_frame(count=12)],
                    backtest_settings(max_trades=2),
                    temp_dir,
                    optimizer_config={
                        "candidate_ranges": {
                            "profit_points": [4, 8],
                            "safety_points": [4, 8],
                            "time_exit": [2],
                            "cooldown": [0, 1],
                            "max_trades": [1, 2],
                        },
                        "refine_top_n": 0,
                        "parallel_workers": 1,
                    },
                    stop_requested=lambda: True,
                )
            self.assertEqual(os.listdir(temp_dir), [])


if __name__ == "__main__":
    unittest.main()

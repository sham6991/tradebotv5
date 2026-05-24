import os
import tempfile
import unittest
from datetime import datetime, timedelta

import openpyxl
import pandas as pd

from live_backtest_optimizer import OPTIMIZED_SETTING_KEYS, run_live_backtest_optimizer, setting_candidates
from tests.test_strategy_regression import settings


class FakeZerodhaHistory:
    def __init__(self):
        self.intervals = []

    def historical_candles(self, instrument_token, from_date, to_date, interval="5minute"):
        self.intervals.append(interval)
        rows = []
        base = datetime.combine(from_date.date(), datetime.min.time()).replace(hour=9, minute=15)
        diffs = [-9, -8, -7, -6, -5, -4, 1, 21, 23, 24, 8, 7, 6, 5, 4, 2, -1, -16, -18, -19]
        rsis = [64, 65, 66, 67, 68, 69, 72, 74, 70, 68, 42, 39, 36, 34, 32, 30, 28, 24, 22, 20]
        for index, diff in enumerate(diffs):
            open_price = 100 + index
            close_price = open_price + 1
            ema50 = 100
            ema20 = ema50 + diff
            volume = 1000 + index
            rows.append({
                "datetime": base + timedelta(minutes=3 * index),
                "open": open_price,
                "high": max(open_price, close_price) + 2,
                "low": min(open_price, close_price) - 2,
                "close": close_price,
                "volume": volume,
                "EMA20": ema20,
                "EMA50": ema50,
                "RSI": rsis[index],
            })
        return pd.DataFrame(rows)


class LiveBacktestOptimizerTests(unittest.TestCase):
    def test_setting_candidates_include_lower_default_and_upper(self):
        self.assertEqual(setting_candidates("rsi_reversal_bullish")[:3], [50, 51, 52])
        self.assertEqual(setting_candidates("rsi_reversal_bearish")[-3:], [48, 49, 50])

    def test_default_optimizer_uses_only_rsi_reversal_settings(self):
        self.assertNotIn("entry_offset", OPTIMIZED_SETTING_KEYS)
        self.assertNotIn("cooldown", OPTIMIZED_SETTING_KEYS)
        self.assertNotIn("profit_points", OPTIMIZED_SETTING_KEYS)
        self.assertEqual(OPTIMIZED_SETTING_KEYS, ["rsi_reversal_bullish", "rsi_reversal_bearish"])

    def test_optimizer_exports_nifty_only_workbook(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp_dir:
            zerodha = FakeZerodhaHistory()
            result = run_live_backtest_optimizer(
                zerodha,
                256265,
                None,
                "2026-05-18",
                "2026-05-18",
                "3minute",
                settings(balance=10000, lot_size=1, max_trades=2, buy_limit_score_low=40),
                temp_dir,
            )

            self.assertTrue(os.path.basename(result["output_path"]).startswith("nifty_optimizer_"))
            self.assertTrue(os.path.exists(result["output_path"]))
            self.assertEqual(result["days_used"], 1)
            self.assertGreater(result["runs"], 1)
            self.assertIn("rsi_reversal_bullish", result["best_settings"])
            self.assertIn("rsi_reversal_bearish", result["best_settings"])
            self.assertEqual(result["best_settings"]["chart_interval"], "3minute")
            self.assertEqual(set(zerodha.intervals), {"3minute"})
            self.assertIn("Bullish RSI Reversal", result["summary"])
            self.assertIn("Bearish RSI Reversal", result["summary"])
            self.assertIn("Bullish Confirmation Rate %", result["summary"])
            self.assertIn("Bearish Confirmation Rate %", result["summary"])

            workbook = openpyxl.load_workbook(result["output_path"], read_only=True)
            self.assertIn("Workbook Guide", workbook.sheetnames)
            self.assertIn("Optimized RSI Values", workbook.sheetnames)
            self.assertIn("Candidate Runs", workbook.sheetnames)
            self.assertIn("Bullish Events", workbook.sheetnames)
            self.assertIn("Bearish Events", workbook.sheetnames)
            self.assertIn("Fetch Log", workbook.sheetnames)

            bullish_headers = [
                cell.value for cell in next(workbook["Bullish Events"].iter_rows(min_row=1, max_row=1))
            ]
            self.assertIn("Setup RSI", bullish_headers)
            self.assertIn("Target EMA Diff", bullish_headers)
            self.assertIn("Confirm Minutes", bullish_headers)

            candidate_headers = [
                cell.value for cell in next(workbook["Candidate Runs"].iter_rows(min_row=1, max_row=1))
            ]
            self.assertIn("Efficiency Score", candidate_headers)
            self.assertIn("Target Cross Rate %", candidate_headers)


if __name__ == "__main__":
    unittest.main()
